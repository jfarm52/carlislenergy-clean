"""Import/Export for bill intake data (CSV)."""

from __future__ import annotations

import csv
import io

from bill_intake.db.connection import get_connection
from bill_intake.db.accounts import upsert_utility_account
from bill_intake.db.meters import upsert_utility_meter
from bill_intake.db.bills_write import insert_bill
from psycopg2.extras import RealDictCursor


def _bill_exists(meter_id, period_start, period_end, total_kwh, total_amount):
    """Check if a bill with the same key fields already exists."""
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            # Match on meter + period + kwh + amount (handles duplicates)
            cur.execute(
                """
                SELECT 1 FROM bills
                WHERE meter_id = %s
                  AND period_start = %s
                  AND period_end = %s
                  AND total_kwh = %s
                  AND total_amount_due = %s
                LIMIT 1
                """,
                (meter_id, period_start, period_end, total_kwh, total_amount),
            )
            return cur.fetchone() is not None
    finally:
        conn.close()


def delete_csv_imported_bills(project_id):
    """
    Delete all CSV-imported bills for a project (bills with bill_file_id = NULL).
    Returns count of deleted bills.
    """
    conn = get_connection()
    try:
        with conn.cursor() as cur:
            # Delete bills that were imported from CSV (no file association)
            cur.execute(
                """
                DELETE FROM bills
                WHERE bill_file_id IS NULL
                  AND account_id IN (SELECT id FROM utility_accounts WHERE project_id = %s)
                """,
                (project_id,),
            )
            deleted = cur.rowcount
            conn.commit()
            print(f"[import] Deleted {deleted} CSV-imported bills for project {project_id}")
            return deleted
    except Exception as e:
        conn.rollback()
        print(f"[import] Error deleting CSV bills: {e}")
        raise
    finally:
        conn.close()


def export_bills_csv(project_id):
    """
    Export all bills for a project as CSV data.
    Returns CSV string with headers and all bill data.
    Format designed for Excel/jMaster import.
    """
    import csv
    import io

    from psycopg2.extras import RealDictCursor

    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    a.utility_name,
                    a.account_number,
                    m.meter_number,
                    b.service_address,
                    b.rate_schedule,
                    b.period_start,
                    b.period_end,
                    b.due_date,
                    b.days_in_period,
                    b.total_kwh,
                    b.total_amount_due,
                    b.blended_rate_dollars,
                    b.avg_cost_per_day,
                    b.energy_charges,
                    b.demand_charges,
                    b.other_charges,
                    b.taxes,
                    b.tou_on_kwh,
                    b.tou_on_rate_dollars,
                    b.tou_on_cost,
                    b.tou_mid_kwh,
                    b.tou_mid_rate_dollars,
                    b.tou_mid_cost,
                    b.tou_off_kwh,
                    b.tou_off_rate_dollars,
                    b.tou_off_cost,
                    b.tou_super_off_kwh,
                    b.tou_super_off_rate_dollars,
                    b.tou_super_off_cost,
                    f.original_filename AS source_file
                FROM bills b
                JOIN utility_accounts a ON b.account_id = a.id
                JOIN utility_meters m ON b.meter_id = m.id
                LEFT JOIN utility_bill_files f ON b.bill_file_id = f.id
                WHERE a.project_id = %s
                ORDER BY a.account_number, m.meter_number, b.period_end DESC
                """,
                (project_id,),
            )
            bills = cur.fetchall()

            if not bills:
                return None

            output = io.StringIO()
            headers = [
                "Utility",
                "Account Number",
                "Meter Number",
                "Service Address",
                "Rate Schedule",
                "Period Start",
                "Period End",
                "Due Date",
                "Days",
                "Total kWh",
                "Total Amount ($)",
                "Blended Rate ($/kWh)",
                "Avg Cost/Day ($)",
                "Energy Charges ($)",
                "Demand Charges ($)",
                "Other Charges ($)",
                "Taxes ($)",
                "On-Peak kWh",
                "On-Peak Rate ($/kWh)",
                "On-Peak Cost ($)",
                "Mid-Peak kWh",
                "Mid-Peak Rate ($/kWh)",
                "Mid-Peak Cost ($)",
                "Off-Peak kWh",
                "Off-Peak Rate ($/kWh)",
                "Off-Peak Cost ($)",
                "Super Off-Peak kWh",
                "Super Off-Peak Rate ($/kWh)",
                "Super Off-Peak Cost ($)",
                "Source File",
            ]

            writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(headers)

            for b in bills:
                row = [
                    b["utility_name"] or "",
                    b["account_number"] or "",
                    b["meter_number"] or "",
                    b["service_address"] or "",
                    b["rate_schedule"] or "",
                    str(b["period_start"]) if b["period_start"] else "",
                    str(b["period_end"]) if b["period_end"] else "",
                    str(b["due_date"]) if b["due_date"] else "",
                    b["days_in_period"] or "",
                    float(b["total_kwh"]) if b["total_kwh"] else "",
                    float(b["total_amount_due"]) if b["total_amount_due"] else "",
                    round(float(b["blended_rate_dollars"]), 4) if b["blended_rate_dollars"] else "",
                    round(float(b["avg_cost_per_day"]), 2) if b["avg_cost_per_day"] else "",
                    float(b["energy_charges"]) if b["energy_charges"] else "",
                    float(b["demand_charges"]) if b["demand_charges"] else "",
                    float(b["other_charges"]) if b["other_charges"] else "",
                    float(b["taxes"]) if b["taxes"] else "",
                    float(b["tou_on_kwh"]) if b["tou_on_kwh"] else "",
                    round(float(b["tou_on_rate_dollars"]), 4) if b["tou_on_rate_dollars"] else "",
                    float(b["tou_on_cost"]) if b["tou_on_cost"] else "",
                    float(b["tou_mid_kwh"]) if b["tou_mid_kwh"] else "",
                    round(float(b["tou_mid_rate_dollars"]), 4) if b["tou_mid_rate_dollars"] else "",
                    float(b["tou_mid_cost"]) if b["tou_mid_cost"] else "",
                    float(b["tou_off_kwh"]) if b["tou_off_kwh"] else "",
                    round(float(b["tou_off_rate_dollars"]), 4) if b["tou_off_rate_dollars"] else "",
                    float(b["tou_off_cost"]) if b["tou_off_cost"] else "",
                    float(b["tou_super_off_kwh"]) if b["tou_super_off_kwh"] else "",
                    round(float(b["tou_super_off_rate_dollars"]), 4)
                    if b["tou_super_off_rate_dollars"]
                    else "",
                    float(b["tou_super_off_cost"]) if b["tou_super_off_cost"] else "",
                    b["source_file"] or "",
                ]
                writer.writerow(row)

            csv_content = output.getvalue()
            output.close()
            return csv_content
    except Exception as e:
        print(f"[bills_db] Error exporting bills CSV: {e}")
        return None
    finally:
        conn.close()


def import_bills_csv(project_id, csv_content):
    """
    Import bills from CSV into a project.
    Creates accounts, meters, and bills from the CSV data.
    Returns dict with import stats.
    
    This is the inverse of export_bills_csv() - allows instant recreation
    of all bill data without re-uploading and re-processing PDFs.
    """
    # CSV header to internal field mapping
    HEADER_MAP = {
        "Utility": "utility_name",
        "Account Number": "account_number",
        "Meter Number": "meter_number",
        "Service Address": "service_address",
        "Rate Schedule": "rate_schedule",
        "Period Start": "period_start",
        "Period End": "period_end",
        "Due Date": "due_date",
        "Days": "days_in_period",
        "Total kWh": "total_kwh",
        "Total Amount ($)": "total_amount_due",
        "Blended Rate ($/kWh)": "blended_rate",
        "Avg Cost/Day ($)": "avg_cost_per_day",
        "Energy Charges ($)": "energy_charges",
        "Demand Charges ($)": "demand_charges",
        "Other Charges ($)": "other_charges",
        "Taxes ($)": "taxes",
        "On-Peak kWh": "tou_on_kwh",
        "On-Peak Rate ($/kWh)": "tou_on_rate_dollars",
        "On-Peak Cost ($)": "tou_on_cost",
        "Mid-Peak kWh": "tou_mid_kwh",
        "Mid-Peak Rate ($/kWh)": "tou_mid_rate_dollars",
        "Mid-Peak Cost ($)": "tou_mid_cost",
        "Off-Peak kWh": "tou_off_kwh",
        "Off-Peak Rate ($/kWh)": "tou_off_rate_dollars",
        "Off-Peak Cost ($)": "tou_off_cost",
        "Super Off-Peak kWh": "tou_super_off_kwh",
        "Super Off-Peak Rate ($/kWh)": "tou_super_off_rate_dollars",
        "Super Off-Peak Cost ($)": "tou_super_off_cost",
        "Source File": "source_file",
    }
    
    def parse_float(val):
        """Parse a float value, returning None for empty/invalid."""
        if val is None or val == "" or val == "None":
            return None
        try:
            return float(str(val).replace(",", "").replace("$", "").strip())
        except (ValueError, TypeError):
            return None
    
    def parse_date(val):
        """Parse a date value, returning None for empty/invalid."""
        if val is None or val == "" or val == "None":
            return None
        val = str(val).strip()
        if not val:
            return None
        # Handle YYYY-MM-DD format from export
        if len(val) == 10 and val[4] == "-":
            return val
        return val
    
    try:
        # Parse CSV
        reader = csv.DictReader(io.StringIO(csv_content))
        
        stats = {
            "accounts_created": 0,
            "meters_created": 0,
            "bills_imported": 0,
            "rows_skipped": 0,
            "errors": []
        }
        
        # Track created entities to count unique creations
        seen_accounts = set()
        seen_meters = set()
        
        for row_num, row in enumerate(reader, start=2):  # Start at 2 (1 is header)
            try:
                # Map headers to internal fields
                data = {}
                for csv_header, internal_key in HEADER_MAP.items():
                    if csv_header in row:
                        data[internal_key] = row[csv_header]
                
                # Required fields check
                utility = data.get("utility_name", "").strip()
                account_num = data.get("account_number", "").strip()
                
                if not utility or not account_num:
                    stats["rows_skipped"] += 1
                    continue
                
                # Get or create account
                account_key = f"{utility}:{account_num}"
                account_id = upsert_utility_account(project_id, utility, account_num)
                if account_key not in seen_accounts:
                    seen_accounts.add(account_key)
                    stats["accounts_created"] += 1
                
                # Get or create meter
                meter_num = data.get("meter_number", "").strip() or "Primary"
                service_addr = data.get("service_address", "").strip()
                meter_key = f"{account_id}:{meter_num}"
                meter_id = upsert_utility_meter(account_id, meter_num, service_addr)
                if meter_key not in seen_meters:
                    seen_meters.add(meter_key)
                    stats["meters_created"] += 1
                
                # Check for duplicate bill (same meter + period + amount)
                period_start = parse_date(data.get("period_start"))
                period_end = parse_date(data.get("period_end"))
                total_kwh = parse_float(data.get("total_kwh"))
                total_amount = parse_float(data.get("total_amount_due"))
                
                if _bill_exists(meter_id, period_start, period_end, total_kwh, total_amount):
                    stats["rows_skipped"] += 1
                    continue
                
                # Insert bill (bill_file_id = None since imported from CSV)
                bill_id = insert_bill(
                    bill_file_id=None,
                    account_id=account_id,
                    meter_id=meter_id,
                    utility_name=utility,
                    service_address=service_addr,
                    rate_schedule=data.get("rate_schedule", "").strip() or None,
                    period_start=parse_date(data.get("period_start")),
                    period_end=parse_date(data.get("period_end")),
                    total_kwh=parse_float(data.get("total_kwh")),
                    total_amount_due=parse_float(data.get("total_amount_due")),
                    energy_charges=parse_float(data.get("energy_charges")),
                    demand_charges=parse_float(data.get("demand_charges")),
                    other_charges=parse_float(data.get("other_charges")),
                    taxes=parse_float(data.get("taxes")),
                    tou_on_kwh=parse_float(data.get("tou_on_kwh")),
                    tou_mid_kwh=parse_float(data.get("tou_mid_kwh")),
                    tou_off_kwh=parse_float(data.get("tou_off_kwh")),
                    tou_super_off_kwh=parse_float(data.get("tou_super_off_kwh")),
                    tou_on_rate_dollars=parse_float(data.get("tou_on_rate_dollars")),
                    tou_mid_rate_dollars=parse_float(data.get("tou_mid_rate_dollars")),
                    tou_off_rate_dollars=parse_float(data.get("tou_off_rate_dollars")),
                    tou_super_off_rate_dollars=parse_float(data.get("tou_super_off_rate_dollars")),
                    tou_on_cost=parse_float(data.get("tou_on_cost")),
                    tou_mid_cost=parse_float(data.get("tou_mid_cost")),
                    tou_off_cost=parse_float(data.get("tou_off_cost")),
                    tou_super_off_cost=parse_float(data.get("tou_super_off_cost")),
                    due_date=parse_date(data.get("due_date")),
                    service_type="electric",
                )
                
                stats["bills_imported"] += 1
                print(f"[import] Row {row_num}: Imported bill {bill_id} for {utility} account {account_num}")
                
            except Exception as e:
                error_msg = f"Row {row_num}: {str(e)}"
                stats["errors"].append(error_msg)
                print(f"[import] ERROR {error_msg}")
        
        print(f"[import] Complete: {stats['bills_imported']} bills, {stats['accounts_created']} accounts, {stats['meters_created']} meters")
        return stats
        
    except Exception as e:
        print(f"[import] ERROR parsing CSV: {e}")
        return {"error": str(e), "bills_imported": 0}


def export_bills_excel(project_id, customer_name=""):
    """
    Export bills as a professionally formatted Excel file for proposals.
    
    Clean, modern design with consistent styling.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    
    conn = get_connection()
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT
                    a.utility_name,
                    a.account_number,
                    m.meter_number,
                    b.service_address,
                    b.period_start,
                    b.period_end,
                    b.days_in_period,
                    b.total_kwh,
                    b.total_amount_due,
                    b.blended_rate_dollars,
                    b.tou_on_kwh,
                    b.tou_on_cost,
                    b.tou_mid_kwh,
                    b.tou_mid_cost,
                    b.tou_off_kwh,
                    b.tou_off_cost,
                    b.tou_super_off_kwh,
                    b.tou_super_off_cost
                FROM bills b
                JOIN utility_accounts a ON b.account_id = a.id
                JOIN utility_meters m ON b.meter_id = m.id
                LEFT JOIN utility_bill_files f ON b.bill_file_id = f.id
                WHERE a.project_id = %s
                  AND b.total_kwh > 0
                ORDER BY b.period_end DESC
                """,
                (project_id,),
            )
            bills = cur.fetchall()
            
            if not bills:
                return None
            
            # Calculate totals
            total_kwh = sum(float(b["total_kwh"] or 0) for b in bills)
            total_cost = sum(float(b["total_amount_due"] or 0) for b in bills)
            total_days = sum(int(b["days_in_period"] or 0) for b in bills)
            avg_rate = (total_cost / total_kwh * 100) if total_kwh > 0 else 0
            avg_daily_kwh = total_kwh / total_days if total_days > 0 else 0
            avg_daily_cost = total_cost / total_days if total_days > 0 else 0
            
            # TOU totals
            tou_on_kwh = sum(float(b["tou_on_kwh"] or 0) for b in bills)
            tou_on_cost = sum(float(b["tou_on_cost"] or 0) for b in bills)
            tou_mid_kwh = sum(float(b["tou_mid_kwh"] or 0) for b in bills)
            tou_mid_cost = sum(float(b["tou_mid_cost"] or 0) for b in bills)
            tou_off_kwh = sum(float(b["tou_off_kwh"] or 0) for b in bills)
            tou_off_cost = sum(float(b["tou_off_cost"] or 0) for b in bills)
            tou_super_off_kwh = sum(float(b["tou_super_off_kwh"] or 0) for b in bills)
            tou_super_off_cost = sum(float(b["tou_super_off_cost"] or 0) for b in bills)
            
            utility_name = bills[0]["utility_name"] if bills else "Unknown"
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Utility Analysis"
            
            # ========== PROFESSIONAL COLOR SCHEME ==========
            # Dark slate blue for headers
            primary_dark = "2C3E50"
            primary_medium = "34495E"
            # Accent color (teal)
            accent = "1ABC9C"
            # Light backgrounds
            light_gray = "ECF0F1"
            white = "FFFFFF"
            # Text colors
            dark_text = "2C3E50"
            
            # Borders
            thin_border = Border(
                left=Side(style='thin', color='BDC3C7'),
                right=Side(style='thin', color='BDC3C7'),
                top=Side(style='thin', color='BDC3C7'),
                bottom=Side(style='thin', color='BDC3C7')
            )
            thick_bottom = Border(bottom=Side(style='medium', color=primary_dark))
            
            # Alignment
            center = Alignment(horizontal='center', vertical='center')
            right_align = Alignment(horizontal='right', vertical='center')
            left_align = Alignment(horizontal='left', vertical='center')
            
            # Set column widths first
            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 16
            ws.column_dimensions['C'].width = 16
            ws.column_dimensions['D'].width = 14
            ws.column_dimensions['E'].width = 16
            ws.column_dimensions['F'].width = 14
            
            # ========== HEADER SECTION ==========
            row = 1
            
            # Main title
            ws.merge_cells('A1:F1')
            ws['A1'] = "UTILITY RATE ANALYSIS"
            ws['A1'].font = Font(name='Arial', size=18, bold=True, color=white)
            ws['A1'].fill = PatternFill(start_color=primary_dark, end_color=primary_dark, fill_type="solid")
            ws['A1'].alignment = center
            ws.row_dimensions[1].height = 35
            
            # Customer name
            ws.merge_cells('A2:F2')
            ws['A2'] = customer_name or "Customer"
            ws['A2'].font = Font(name='Arial', size=14, bold=True, color=dark_text)
            ws['A2'].fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type="solid")
            ws['A2'].alignment = center
            ws.row_dimensions[2].height = 25
            
            # Utility name
            ws.merge_cells('A3:F3')
            ws['A3'] = utility_name
            ws['A3'].font = Font(name='Arial', size=11, italic=True, color=primary_medium)
            ws['A3'].fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type="solid")
            ws['A3'].alignment = center
            ws.row_dimensions[3].height = 22
            
            # Spacer row
            ws.row_dimensions[4].height = 10
            
            # ========== SUMMARY METRICS ==========
            row = 5
            
            # Summary header
            ws.merge_cells('A5:F5')
            ws['A5'] = "ANNUAL SUMMARY"
            ws['A5'].font = Font(name='Arial', size=12, bold=True, color=white)
            ws['A5'].fill = PatternFill(start_color=accent, end_color=accent, fill_type="solid")
            ws['A5'].alignment = center
            ws.row_dimensions[5].height = 25
            
            # Metric labels row
            row = 6
            metrics = [
                ("Total Usage", f"{total_kwh:,.0f} kWh"),
                ("Total Cost", f"${total_cost:,.2f}"),
                ("Avg Rate", f"{avg_rate:.1f}¢/kWh"),
            ]
            metrics2 = [
                ("Billing Periods", f"{len(bills)}"),
                ("Daily Avg", f"{avg_daily_kwh:,.0f} kWh"),
                ("Daily Cost", f"${avg_daily_cost:,.2f}"),
            ]
            
            # Row 6 - labels
            for i, (label, _) in enumerate(metrics):
                col = get_column_letter(i * 2 + 1)
                ws.merge_cells(f'{col}6:{get_column_letter(i * 2 + 2)}6')
                ws[f'{col}6'] = label
                ws[f'{col}6'].font = Font(name='Arial', size=9, color=primary_medium)
                ws[f'{col}6'].alignment = center
            
            # Row 7 - values
            for i, (_, value) in enumerate(metrics):
                col = get_column_letter(i * 2 + 1)
                ws.merge_cells(f'{col}7:{get_column_letter(i * 2 + 2)}7')
                ws[f'{col}7'] = value
                ws[f'{col}7'].font = Font(name='Arial', size=14, bold=True, color=dark_text)
                ws[f'{col}7'].alignment = center
            ws.row_dimensions[7].height = 28
            
            # Row 8 - second row labels
            for i, (label, _) in enumerate(metrics2):
                col = get_column_letter(i * 2 + 1)
                ws.merge_cells(f'{col}8:{get_column_letter(i * 2 + 2)}8')
                ws[f'{col}8'] = label
                ws[f'{col}8'].font = Font(name='Arial', size=9, color=primary_medium)
                ws[f'{col}8'].alignment = center
            
            # Row 9 - second row values
            for i, (_, value) in enumerate(metrics2):
                col = get_column_letter(i * 2 + 1)
                ws.merge_cells(f'{col}9:{get_column_letter(i * 2 + 2)}9')
                ws[f'{col}9'] = value
                ws[f'{col}9'].font = Font(name='Arial', size=12, bold=True, color=dark_text)
                ws[f'{col}9'].alignment = center
            
            # Spacer
            ws.row_dimensions[10].height = 15
            
            # ========== TIME-OF-USE BREAKDOWN ==========
            row = 11
            
            ws.merge_cells('A11:F11')
            ws['A11'] = "TIME-OF-USE BREAKDOWN"
            ws['A11'].font = Font(name='Arial', size=12, bold=True, color=white)
            ws['A11'].fill = PatternFill(start_color=primary_medium, end_color=primary_medium, fill_type="solid")
            ws['A11'].alignment = center
            ws.row_dimensions[11].height = 25
            
            # TOU headers - merge E and F for cleaner look
            row = 12
            tou_headers = ["Period", "Usage (kWh)", "Cost", "Rate (¢/kWh)", "% of Total"]
            for i, header in enumerate(tou_headers):
                col = get_column_letter(i + 1)
                if i == 4:  # % of Total - merge E and F
                    ws.merge_cells('E12:F12')
                ws[f'{col}12'] = header
                ws[f'{col}12'].font = Font(name='Arial', size=10, bold=True, color=dark_text)
                ws[f'{col}12'].fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type="solid")
                ws[f'{col}12'].alignment = center
                ws[f'{col}12'].border = thin_border
            ws['F12'].border = thin_border  # Right border on merged cell
            
            # TOU data with colored header cells
            tou_data = []
            if tou_on_kwh > 0:
                tou_data.append(("On-Peak", tou_on_kwh, tou_on_cost, "C0392B"))  # Dark red
            if tou_mid_kwh > 0:
                tou_data.append(("Mid-Peak", tou_mid_kwh, tou_mid_cost, "D68910"))  # Dark orange
            if tou_off_kwh > 0:
                tou_data.append(("Off-Peak", tou_off_kwh, tou_off_cost, "1E8449"))  # Dark green
            if tou_super_off_kwh > 0:
                tou_data.append(("Super Off-Peak", tou_super_off_kwh, tou_super_off_cost, "2874A6"))  # Dark blue
            
            row = 13
            for period, kwh, cost, color in tou_data:
                rate = (cost / kwh * 100) if kwh > 0 else 0
                pct = (kwh / total_kwh * 100) if total_kwh > 0 else 0
                
                # Period name - filled cell with white text, centered
                ws[f'A{row}'] = period
                ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True, color=white)
                ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                ws[f'A{row}'].alignment = center
                ws[f'A{row}'].border = thin_border
                
                ws[f'B{row}'] = kwh
                ws[f'B{row}'].number_format = '#,##0'
                ws[f'B{row}'].alignment = right_align
                ws[f'B{row}'].border = thin_border
                
                ws[f'C{row}'] = cost
                ws[f'C{row}'].number_format = '"$"#,##0.00'
                ws[f'C{row}'].alignment = right_align
                ws[f'C{row}'].border = thin_border
                
                ws[f'D{row}'] = rate
                ws[f'D{row}'].number_format = '0.0'
                ws[f'D{row}'].alignment = right_align
                ws[f'D{row}'].border = thin_border
                
                # % of Total - merge E and F, centered
                ws.merge_cells(f'E{row}:F{row}')
                ws[f'E{row}'] = pct / 100
                ws[f'E{row}'].number_format = '0.0%'
                ws[f'E{row}'].alignment = center
                ws[f'E{row}'].border = thin_border
                ws[f'F{row}'].border = thin_border  # Right border
                
                row += 1
            
            # Spacer
            ws.row_dimensions[row].height = 15
            row += 1
            
            # ========== BILLING HISTORY ==========
            history_start = row
            
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = "BILLING HISTORY"
            ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True, color=white)
            ws[f'A{row}'].fill = PatternFill(start_color=primary_medium, end_color=primary_medium, fill_type="solid")
            ws[f'A{row}'].alignment = center
            ws.row_dimensions[row].height = 25
            row += 1
            
            # History headers
            hist_headers = ["Month", "Usage (kWh)", "Charge", "Rate (¢/kWh)", "Days", "$/Day"]
            for i, header in enumerate(hist_headers):
                col = get_column_letter(i + 1)
                ws[f'{col}{row}'] = header
                ws[f'{col}{row}'].font = Font(name='Arial', size=10, bold=True, color=dark_text)
                ws[f'{col}{row}'].fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type="solid")
                ws[f'{col}{row}'].alignment = center
                ws[f'{col}{row}'].border = thin_border
            row += 1
            
            # History data with alternating rows
            for idx, b in enumerate(bills):
                period_end = b["period_end"]
                month_str = period_end.strftime("%b %Y") if hasattr(period_end, 'strftime') else str(period_end)[:7]
                kwh_val = float(b["total_kwh"] or 0)
                cost_val = float(b["total_amount_due"] or 0)
                days_val = int(b["days_in_period"] or 30)
                rate_val = (cost_val / kwh_val * 100) if kwh_val > 0 else 0
                daily_cost = cost_val / days_val if days_val > 0 else 0
                
                # Alternating row color
                row_fill = PatternFill(start_color=white, end_color=white, fill_type="solid") if idx % 2 == 0 else PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
                
                ws[f'A{row}'] = month_str
                ws[f'A{row}'].alignment = center
                ws[f'A{row}'].border = thin_border
                ws[f'A{row}'].fill = row_fill
                
                ws[f'B{row}'] = kwh_val
                ws[f'B{row}'].number_format = '#,##0'
                ws[f'B{row}'].alignment = right_align
                ws[f'B{row}'].border = thin_border
                ws[f'B{row}'].fill = row_fill
                
                ws[f'C{row}'] = cost_val
                ws[f'C{row}'].number_format = '"$"#,##0.00'
                ws[f'C{row}'].alignment = right_align
                ws[f'C{row}'].border = thin_border
                ws[f'C{row}'].fill = row_fill
                
                ws[f'D{row}'] = rate_val
                ws[f'D{row}'].number_format = '0.0'
                ws[f'D{row}'].alignment = right_align
                ws[f'D{row}'].border = thin_border
                ws[f'D{row}'].fill = row_fill
                
                ws[f'E{row}'] = days_val
                ws[f'E{row}'].alignment = right_align
                ws[f'E{row}'].border = thin_border
                ws[f'E{row}'].fill = row_fill
                
                ws[f'F{row}'] = daily_cost
                ws[f'F{row}'].number_format = '"$"#,##0.00'
                ws[f'F{row}'].alignment = right_align
                ws[f'F{row}'].border = thin_border
                ws[f'F{row}'].fill = row_fill
                
                row += 1
            
            # Totals row
            ws[f'A{row}'] = "TOTAL"
            ws[f'A{row}'].font = Font(name='Arial', size=10, bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type="solid")
            ws[f'A{row}'].alignment = center
            ws[f'A{row}'].border = thin_border
            
            ws[f'B{row}'] = total_kwh
            ws[f'B{row}'].number_format = '#,##0'
            ws[f'B{row}'].font = Font(bold=True)
            ws[f'B{row}'].fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type="solid")
            ws[f'B{row}'].alignment = right_align
            ws[f'B{row}'].border = thin_border
            
            ws[f'C{row}'] = total_cost
            ws[f'C{row}'].number_format = '"$"#,##0.00'
            ws[f'C{row}'].font = Font(bold=True)
            ws[f'C{row}'].fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type="solid")
            ws[f'C{row}'].alignment = right_align
            ws[f'C{row}'].border = thin_border
            
            ws[f'D{row}'] = avg_rate
            ws[f'D{row}'].number_format = '0.0'
            ws[f'D{row}'].font = Font(bold=True)
            ws[f'D{row}'].fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type="solid")
            ws[f'D{row}'].alignment = right_align
            ws[f'D{row}'].border = thin_border
            
            ws[f'E{row}'] = total_days
            ws[f'E{row}'].font = Font(bold=True)
            ws[f'E{row}'].fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type="solid")
            ws[f'E{row}'].alignment = right_align
            ws[f'E{row}'].border = thin_border
            
            ws[f'F{row}'] = avg_daily_cost
            ws[f'F{row}'].number_format = '"$"#,##0.00'
            ws[f'F{row}'].font = Font(bold=True)
            ws[f'F{row}'].fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type="solid")
            ws[f'F{row}'].alignment = right_align
            ws[f'F{row}'].border = thin_border
            
            # ========== RAW DATA (Hidden columns for formulas) ==========
            raw_start_col = 27  # Column AA
            raw_headers = [
                "Month", "Period Start", "Period End", "Days", "kWh", "Charge", "Rate",
                "On-Peak kWh", "On-Peak Cost", "Mid-Peak kWh", "Mid-Peak Cost",
                "Off-Peak kWh", "Off-Peak Cost", "Super Off kWh", "Super Off Cost"
            ]
            for i, header in enumerate(raw_headers):
                col = get_column_letter(raw_start_col + i)
                ws[f'{col}1'] = header
                ws[f'{col}1'].font = Font(bold=True, size=9)
            
            for row_idx, b in enumerate(bills, start=2):
                period_end = b["period_end"]
                month_str = period_end.strftime("%b %Y") if hasattr(period_end, 'strftime') else str(period_end)
                raw_data = [
                    month_str,
                    str(b["period_start"]) if b["period_start"] else "",
                    str(b["period_end"]) if b["period_end"] else "",
                    b["days_in_period"] or 0,
                    float(b["total_kwh"] or 0),
                    float(b["total_amount_due"] or 0),
                    float(b["blended_rate_dollars"] or 0) * 100,
                    float(b["tou_on_kwh"] or 0),
                    float(b["tou_on_cost"] or 0),
                    float(b["tou_mid_kwh"] or 0),
                    float(b["tou_mid_cost"] or 0),
                    float(b["tou_off_kwh"] or 0),
                    float(b["tou_off_cost"] or 0),
                    float(b["tou_super_off_kwh"] or 0),
                    float(b["tou_super_off_cost"] or 0),
                ]
                for i, val in enumerate(raw_data):
                    col = get_column_letter(raw_start_col + i)
                    ws[f'{col}{row_idx}'] = val
            
            # Print settings
            ws.print_title_rows = '1:3'
            ws.page_setup.orientation = 'portrait'
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
            
    except Exception as e:
        print(f"[export] Error generating Excel: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        conn.close()
